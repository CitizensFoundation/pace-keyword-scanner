import string
import pandas as pd
import pathlib
import os
import pprint
from openpyxl import load_workbook
import sys
import re
from imblearn.over_sampling import SMOTE

TO_LOWER = True

class XlsManager:
    topics = {}
    language = None
    inFileNames = []
    inFilesPath = None
    outFilesPath = None
    outFilePath = None
    fromXlsDataRounds = []
    toXlsWorkbook = None
    keywordsSheet = None
    xCount = 0
    notXCount = 0
    inconsistentRatingCount = 0
    itemsToRate = []
    texts = []
    relevantTexts = []
    notRelevantTexts = []
    insconsistencies = []


    # default constructor
    def __init__(self, language):
        self.language = language
        currentPath = pathlib.Path().absolute()
        self.inFilesPath = f"{currentPath}/xls/fromXls/{self.language}/"
        self.inFileNames = sorted(os.listdir(self.inFilesPath))
        self.outFilesPath = f"{currentPath}/xls/toXls/{self.language}/"
        self.outFileNames = sorted(os.listdir(self.outFilesPath))
        print(self.outFilesPath )

    def remove_punct(self, text):
        table=str.maketrans('','',string.punctuation)
        return text.translate(table)

    def cleanup_text(self, sentence):
        #return sentence
        sentence = re.sub('[^a-zA-Z]', ' ', sentence)

        # Single character removal
        #sentence = re.sub(r"\s+[a-zA-Z]\s+", ' ', sentence)

        # Removing multiple spaces
        sentence = re.sub(r'\s+', ' ', sentence)

        return sentence

    def setup_all_from_xls(self):
        for fileName in self.inFileNames:
            round = []
            xlsx = pd.ExcelFile(f"{self.inFilesPath}/{fileName}")
            print(fileName)
            for sheet in xlsx.sheet_names:
                round.append(xlsx.parse(sheet))
            self.fromXlsDataRounds.append(round)
            #print(round)
        self.keywordsSheet = self.fromXlsDataRounds[-1][0]
        print(f"LEN ROUNDS {len(self.fromXlsDataRounds)}")
        self.read_in_topics()

        if len(self.outFileNames)>1:
            print(f"Too many files in the {self.outFileNames} folder")
            sys.exit()
        elif len(self.outFileNames)==1:
            self.outFilePath = f"{self.outFilesPath}{self.outFileNames[0]}"
            print(f"Loading out xls from {self.outFilePath}")
            self.toXlsWorkbook = load_workbook(self.outFilePath)
            print(f"Loaded {self.toXlsWorkbook}")

    def get_items_to_rate(self):
        self.itemsToRate = []
        index = 0
        print(f"Do I have {self.toXlsWorkbook}")
        for worksheet in self.toXlsWorkbook.worksheets:
            if index>0:
                rowIndex = 0
                for row in worksheet.iter_rows():
                    subTopic = None
                    text = None
                    rateAble = False
                    if rowIndex>1 and row[0].value and row[1].value:
                        subTopic = row[0].value.strip()
                        if TO_LOWER:
                          text = row[1].value.strip().lower()
                        else:
                          text = row[1].value.strip()
                        text = self.cleanup_text(text)
                        rateAble = True

                    newItem = {
                        "topic": worksheet.title,
                        "sheet": worksheet,
                        "row": row,
                        "index": rowIndex,
                        "ratingCol": row[2],
                        "text": text,
                        "subTopic": subTopic,
                        "rateAble": rateAble
                        }
                    #print(newItem)
                    self.itemsToRate.append(newItem)
                    rowIndex+=1
            index+=1
        #return self.itemsToRate[:2000]
        return self.itemsToRate

    def save_predicted_items(self):
        print(f"Saving {self.outFilePath}")
        self.toXlsWorkbook.save(self.outFilePath)

    def read_in_topics(self):
        topicNames = []
        for index, row in self.keywordsSheet.iterrows():
            if (index>1 and not pd.isnull(row[2]) and not pd.isnull(row[3])):
                topic = row[2]
                subTopic = row[3]
                if (not topic in self.topics):
                    self.topics[topic] = {}
                elif (not subTopic in self.topics[topic]):
                   self.topics[topic][subTopic] = subTopic

    def skip_sheet(self, sheet, options):
        sheetTopic = sheet.head(0).columns[0]
        sheetTopic = sheetTopic.strip()
        #print(f"SheetTopic {sheetTopic}")
        if options.get("topic"):
            if sheetTopic.lower()==options["topic"].lower():
                return False
            else:
                return True
        elif options.get('onlyTopics'):
            return sheetTopic not in options.get('onlyTopics')
        elif options.get('skipTopics'):
            return sheetTopic in options.get('skipTopics')
        else:
          return False

    def check_inconsistent_ratings(self, text, rating):
        if rating=="1":
            self.relevantTexts.append(text)
            if text in self.notRelevantTexts:
                self.inconsistentRatingCount+=1
                return True
            else:
                return False
        elif rating=="x":
            self.notRelevantTexts.append(text)
            if text in self.relevantTexts:
                self.inconsistentRatingCount+=1
                return True
            else:
                return False
        else:
            print(f"ERROR UNKNOWN RATING {rating}")

    def add_out_data_from_row(self, sheetName, rowIndex, row, outData, options):
        subTopic = row[0].strip()
        text = row[1]

        dedupTexts = options.get('dedupTexts')==True

        if isinstance(text, str):
            if TO_LOWER:
                text = row[1].strip().lower()
            else:
                text = row[1].strip()

            text = self.cleanup_text(text)
            text = text.strip()
            #print(f"{text}")

            rating = row[2]

            if text!="x" and options.get('allowEmptyRatings'):
                if not rating or rating=="" or rating==" " or pd.isnull(rating):
                    rating = "x"
                else:
                    rating = "1"

                if self.check_inconsistent_ratings(text, rating):
                    reversedRating = " " if rating=="x" else "x"
                    self.insconsistencies.append(f"INCONSISTENT RATING FOR: {text[:140]} - {reversedRating} {rowIndex}")

            if text!="x" and not dedupTexts or text not in self.texts:
                self.texts.append(text)

                if isinstance(rating, str):
                    rating = rating.strip().lower()

                if options.get("subTopic") and options.get("subTopic")!=subTopic:
                    return

                if options.get('onlyOnes'):
                    if isinstance(rating, int) and rating==1:
                        outData.append([text, 1])
                        self.notXCount += 1
                    else:
                        outData.append([text, 0])
                        self.xCount += 1
                elif options.get('onlyOneTwo'):
                    if isinstance(rating, int) and (rating==1 or rating==2):
                        outData.append([text, 1])
                        self.notXCount += 1
                    elif rating==3 or rating=="x":
                        outData.append([text, 0])
                        self.xCount += 1
                else:
                    if rating=="x" and not options.get("trainOnlyRelevant"):
                        self.xCount += 1
                        outData.append([text, 0])
                    elif not rating=="x":
                        self.notXCount += 1
                        if options.get("trainOnlyRelevant"):
                            if isinstance(rating, int):
                                rating -= 1
                                if rating==0 or rating==1 or rating==2:
                                    outData.append([text, rating])
                                else:
                                    print(f"Wrong rating value {rating}")
                            else:
                                print(f"Wrong rating format {rating}")
                        else:
                            outData.append([text, 1])
            #else:
             #   print(f"Duplicate text")
        else:
            print(f"WARNING: Found non text element in training data {subTopic} text {row[1]} rating {row[2]}")

    def get_training_data(self, options):
        outData = []
        self.xCount = 0
        self.notXCount = 0
        #print(self.fromXlsDataRounds[0]);
        for round in self.fromXlsDataRounds:
            #print(round)
            first = True
            for sheet in round:
                # Skip the first keyword sheets
                if not first:
                    #print(sheet)
                    if not self.skip_sheet(sheet, options):
                        for index, row in sheet.iterrows():
                            if index>1 and not pd.isnull(row[0]) and not pd.isnull(row[1]) and (options.get("allowEmptyRatings")==True or not pd.isnull(row[2])):
                                self.add_out_data_from_row(sheet,index,row,outData,options)
                else:
                    first = False

        if options.get('onlyOnes'):
            print(f"Not one coded: {self.xCount}")
            print(f"One coded: {self.notXCount}")
        elif options.get('onlyOneTwo'):
            print(f"Not oneTwo coded: {self.xCount}")
            print(f"OneTwo coded: {self.notXCount}")
        else:
            print(f"x coded: {self.xCount}")
            print(f"Not x coded: {self.notXCount}")
            print(f"Inconsistent rating count: {self.inconsistentRatingCount}")

            for inconsistency in sorted(self.insconsistencies):
                print(inconsistency)
        return outData

    def get_balanced_training_data(self, options):
        unbalancedOutData = self.get_training_data(options)
        oneCodedTexsts = []
        xCodedTexts = []
        outData = []

        for data in unbalancedOutData:
            if data[1]==1:
                oneCodedTexsts.append(data)
            else:
                xCodedTexts.append(data)

        oneCodedResampled, xCodedResampled = SMOTE().fit_resample(oneCodedTexsts, xCodedTexts)

        for oneCodedText in oneCodedResampled:
            outData.append([oneCodedText, 1])

        for xCodedText in xCodedResampled:
            outData.append([xCodedText, 0])

        print(f"Resampled 0 coded length: {len(xCodedResampled)}")
        print(f"Resampled 1 coded length: {len(oneCodedResampled)}")
        return outData


    def example_iterate_over_topis(self):
        for topic in self.topics:
            print(topic)
            for subTopic in self.topics.get(topic):
                print(f"     {subTopic}")

    def basic_test(self, options):
        self.setup_all_from_xls()
        trainingData = self.get_training_data(options)

        relevant = []
        notRelevant= []

        for entry in trainingData:
            if (entry[1]==1):
                relevant.append(entry)
            else:
                notRelevant.append(entry)

        print(len(trainingData))
        print(len(relevant))
        print(len(notRelevant))


